"""
carga_ons.py
============
Download automático da carga efetiva diária (MWmed) do ONS/SDRO para os dois
últimos meses completos e salva:
  - Carga_Energia_ONS_AAAA_MM.xlsx  (um arquivo por mês)
  - Carga_Energia_ONS_Consolidado.xlsx  (série completa + gráfico)

Fonte: http://sdro.ons.org.br/SDRO/DIARIO/
Requisitos: pip install requests beautifulsoup4 openpyxl
"""

import calendar
import datetime
import re
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
BASE_URL = "https://sdro.ons.org.br/SDRO/DIARIO/{date}/HTML/15_CargaDiariaSubmercado.html"
SUBSISTEMAS = ["Sudeste/Centro-Oeste", "Sul", "Nordeste", "Norte", "Sistema Interligado Nacional"]
HEADERS = {"User-Agent": "Mozilla/5.0"}


def _dois_ultimos_meses():
    """Retorna lista de (ano, mes) para os dois últimos meses completos."""
    hoje = datetime.date.today()
    meses = []
    for delta in range(2, 0, -1):
        d = hoje.replace(day=1) - datetime.timedelta(days=delta * 28)
        d = d.replace(day=1)
        meses.append((d.year, d.month))
    # Garante ordem crescente e unicidade
    meses = sorted(set(meses))
    return meses


def _datas_do_mes(ano, mes):
    _, ultimo_dia = calendar.monthrange(ano, mes)
    return [datetime.date(ano, mes, dia) for dia in range(1, ultimo_dia + 1)]


# ---------------------------------------------------------------------------
# Download e parsing
# ---------------------------------------------------------------------------
def _parse_mwmed(html: str, data: datetime.date) -> dict | None:
    """Extrai a tabela de MWmed do HTML. Retorna dict {subsistema: valor} ou None."""
    soup = BeautifulSoup(html, "html.parser")
    tabelas = soup.find_all("table")

    for tabela in tabelas:
        linhas = tabela.find_all("tr")
        # Identifica a tabela de MWmed pelo cabeçalho
        cabecalhos = [th.get_text(strip=True) for th in linhas[0].find_all(["th", "td"])] if linhas else []
        if not any("MWmed" in c for c in cabecalhos):
            continue

        resultado = {"Data": data}
        for linha in linhas[1:]:
            colunas = [td.get_text(strip=True) for td in linha.find_all(["td", "th"])]
            if len(colunas) < 2:
                continue
            nome = colunas[0]
            if nome in SUBSISTEMAS:
                try:
                    valor = float(colunas[1].replace(".", "").replace(",", "."))
                except ValueError:
                    valor = None
                resultado[nome] = valor
        if len(resultado) > 1:
            return resultado
    return None


def baixar_mes(ano: int, mes: int, delay: float = 0.3) -> list[dict]:
    """Baixa todos os dias de um mês e retorna lista de registros."""
    registros = []
    for data in _datas_do_mes(ano, mes):
        url = BASE_URL.format(date=data.strftime("%Y_%m_%d"))
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                reg = _parse_mwmed(r.text, data)
                if reg:
                    registros.append(reg)
                    print(f"  ✓ {data.isoformat()}: SIN = {reg.get('Sistema Interligado Nacional'):,.2f} MWmed")
                else:
                    print(f"  ✗ {data.isoformat()}: sem dados")
            else:
                print(f"  ✗ {data.isoformat()}: HTTP {r.status_code}")
        except Exception as e:
            print(f"  ✗ {data.isoformat()}: {e}")
        time.sleep(delay)
    return registros


# ---------------------------------------------------------------------------
# Formatação / estilos
# ---------------------------------------------------------------------------
AZUL_ESCURO = "1F3864"
AZUL_CLARO  = "D6E4F0"
CINZA_LINHA = "F2F2F2"

def _estilo_cabecalho(ws, linha: int, n_colunas: int):
    fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col in range(1, n_colunas + 1):
        cell = ws.cell(row=linha, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _borda_fina():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _formatar_dados(ws, linha_inicio: int, n_linhas: int, n_colunas: int):
    borda = _borda_fina()
    for i in range(n_linhas):
        fill = PatternFill("solid", fgColor=CINZA_LINHA if i % 2 == 0 else "FFFFFF")
        for col in range(1, n_colunas + 1):
            cell = ws.cell(row=linha_inicio + i, column=col)
            cell.fill = fill
            cell.border = borda
            cell.font = Font(name="Arial", size=9)
            if col > 1:  # colunas numéricas
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Criação dos Excel
# ---------------------------------------------------------------------------
COLUNAS = ["Data"] + SUBSISTEMAS


def _criar_aba(wb: Workbook, registros: list[dict], nome_aba: str):
    ws = wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False

    # Cabeçalho
    for col, titulo in enumerate(COLUNAS, start=1):
        ws.cell(row=1, column=col, value=titulo)
    _estilo_cabecalho(ws, 1, len(COLUNAS))

    # Dados
    for i, reg in enumerate(registros, start=2):
        for col, campo in enumerate(COLUNAS, start=1):
            ws.cell(row=i, column=col, value=reg.get(campo))

    _formatar_dados(ws, 2, len(registros), len(COLUNAS))

    # Larguras de coluna
    ws.column_dimensions["A"].width = 13
    for col in range(2, len(COLUNAS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Linha de totais/médias
    linha_media = len(registros) + 2
    ws.cell(row=linha_media, column=1, value="MÉDIA DO MÊS")
    ws.cell(row=linha_media, column=1).font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=linha_media, column=1).fill = PatternFill("solid", fgColor=AZUL_CLARO)
    for col in range(2, len(COLUNAS) + 1):
        letra = get_column_letter(col)
        ws.cell(row=linha_media, column=col,
                value=f"=AVERAGE({letra}2:{letra}{len(registros)+1})")
        ws.cell(row=linha_media, column=col).number_format = "#,##0.00"
        ws.cell(row=linha_media, column=col).font = Font(bold=True, name="Arial", size=9)
        ws.cell(row=linha_media, column=col).fill = PatternFill("solid", fgColor=AZUL_CLARO)

    # Rodapé com fonte
    ws.cell(row=linha_media + 2, column=1,
            value="Fonte: ONS – Boletim Diário da Operação (sdro.ons.org.br)")
    ws.cell(row=linha_media + 2, column=1).font = Font(italic=True, size=8, color="808080", name="Arial")
    ws.merge_cells(start_row=linha_media + 2, start_column=1,
                   end_row=linha_media + 2, end_column=len(COLUNAS))

    return ws


def _adicionar_grafico(ws, n_linhas: int):
    """Adiciona gráfico de linha do SIN (coluna F = col 6) na aba."""
    chart = LineChart()
    chart.title = "Carga Efetiva Diária – SIN (MWmed)"
    chart.style = 10
    chart.y_axis.title = "MWmed"
    chart.x_axis.title = "Data"
    chart.height = 14
    chart.width = 28

    # Série: coluna 6 (Sistema Interligado Nacional)
    dados = Reference(ws, min_col=6, min_row=1, max_row=n_linhas + 1)
    rotulos = Reference(ws, min_col=1, min_row=2, max_row=n_linhas + 1)
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(rotulos)
    chart.series[0].graphicalProperties.line.solidFill = "1F78B4"

    ws.add_chart(chart, f"A{n_linhas + 5}")


def salvar_por_mes(registros: list[dict], ano: int, mes: int) -> str:
    nome_mes = datetime.date(ano, mes, 1).strftime("%B").capitalize()
    nome_arquivo = f"Carga_Energia_ONS_{ano}_{mes:02d}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão
    ws = _criar_aba(wb, registros, f"{nome_mes} {ano}")
    _adicionar_grafico(ws, len(registros))

    wb.save(nome_arquivo)
    print(f"\n  → Salvo: {nome_arquivo}")
    return nome_arquivo


def salvar_consolidado(todos: list[dict]) -> str:
    nome_arquivo = "Carga_Energia_ONS_Consolidado.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    # Aba consolidada
    ws = _criar_aba(wb, todos, "Série Histórica")
    _adicionar_grafico(ws, len(todos))

    # Abas mensais separadas
    from itertools import groupby
    for (ano, mes), grupo in groupby(todos, key=lambda r: (r["Data"].year, r["Data"].month)):
        nome_mes = datetime.date(ano, mes, 1).strftime("%b/%Y")
        _criar_aba(wb, list(grupo), nome_mes)

    wb.save(nome_arquivo)
    print(f"  → Salvo: {nome_arquivo}")
    return nome_arquivo


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    meses = _dois_ultimos_meses()
    print(f"Baixando dados para: {[f'{m[0]}-{m[1]:02d}' for m in meses]}\n")

    todos = []
    arquivos_mes = []

    for ano, mes in meses:
        nome_mes = datetime.date(ano, mes, 1).strftime("%B %Y")
        print(f"=== {nome_mes} ===")
        registros = baixar_mes(ano, mes)
        if registros:
            arq = salvar_por_mes(registros, ano, mes)
            arquivos_mes.append(arq)
            todos.extend(registros)

    if todos:
        print("\n=== Consolidado ===")
        salvar_consolidado(todos)
        print(f"\nConcluído: {len(todos)} dias baixados.")
        print("Arquivos gerados:")
        for arq in arquivos_mes:
            print(f"  • {arq}")
        print("  • Carga_Energia_ONS_Consolidado.xlsx")
    else:
        print("Nenhum dado encontrado.")


if __name__ == "__main__":
    main()
